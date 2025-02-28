from fractions import Fraction

from openpyxl import *

from createListsInKeep import createListsInKeep

wb = load_workbook(filename="/home/rammas/Downloads/Menu.xlsx")

where_ingredients = {}

sheet_where_ingredientes = wb["Enums"]

start = 3
ingredients_col = "D"
where_col = "E"
product = sheet_where_ingredientes[f"{ingredients_col}{start}"].value
where = sheet_where_ingredientes[f"{where_col}{start}"].value
while product != None:
    where_ingredients[product] = where
    start += 1
    product = sheet_where_ingredientes[f"{ingredients_col}{start}"].value
    where = sheet_where_ingredientes[f"{where_col}{start}"].value

sheet_ingredientes = wb["Ingredientes"]

col_cantidad = "B"
col_unidad = "C"
col_producto = "D"
cell = 2

master_ingredients = {}
value = sheet_ingredientes[f"{col_cantidad}{cell}"].value
while value != None and sheet_ingredientes[f"{col_cantidad}{cell + 1}"].value != None:
    nombre = value
    master_ingredients[nombre] = []
    cell += 1
    for recipe_rows in range(19):
        cell += 1
        value = sheet_ingredientes[f"{col_cantidad}{cell}"].value
        if value != None:
            cantidad = str(sheet_ingredientes[f"{col_cantidad}{cell}"].value)
            cantidad = cantidad.replace("=", "")
            unidad = sheet_ingredientes[f"{col_unidad}{cell}"].value
            producto = sheet_ingredientes[f"{col_producto}{cell}"].value
            master_ingredients[nombre].append(
                {"cantidad": cantidad, "unidad": unidad, "producto": producto}
            )

    cell += 1
    value = sheet_ingredientes[f"{col_cantidad}{cell}"].value

sheet_menu = wb["1_Menú"]

shopping_list = {}

print("Select one Menu:")
print("(1) Top (Green)")
print("(2) Bottom (Magenta)")
print("(3) Both")
menu_selected = input()
menu = int(menu_selected)
if menu < 1 or menu > 3:
    print(f"Out of range: {menu}")
    exit

menu_row_range = range(3, 8) if (menu == 1 or menu == 3) else range(11, 16)
for menu_col in range(66, 73):  # 'B' to 'H'
    for menu_row in menu_row_range:
        value = sheet_menu[f"{chr(menu_col)}{menu_row}"].value

        if not value:
            continue

        ingredientes = master_ingredients[value]
        None
        for ingrediente in ingredientes:
            if not ingrediente["producto"] in shopping_list:
                shopping_list[ingrediente["producto"]] = {
                    ingrediente["unidad"]: ingrediente["cantidad"]
                }

            else:
                if ingrediente["unidad"] in shopping_list[ingrediente["producto"]]:
                    cantidad = str(ingrediente["cantidad"])
                    cantidad = float(sum(Fraction(s) for s in cantidad.split()))
                    cantidad_anterior = str(
                        shopping_list[ingrediente["producto"]][ingrediente["unidad"]]
                    )
                    cantidad_anterior = float(
                        sum(Fraction(s) for s in cantidad_anterior.split())
                    )
                    cantidad += cantidad_anterior

                    shopping_list[ingrediente["producto"]][
                        ingrediente["unidad"]
                    ] = cantidad

                else:
                    shopping_list[ingrediente["producto"]][ingrediente["unidad"]] = (
                        ingrediente["cantidad"]
                    )

if menu == 3:
    menu_row_range = range(11, 16)
    for menu_col in range(66, 73):  # 'B' to 'H'
        for menu_row in menu_row_range:
            value = sheet_menu[f"{chr(menu_col)}{menu_row}"].value

            if not value:
                continue

            ingredientes = master_ingredients[value]
            for ingrediente in ingredientes:
                if not ingrediente["producto"] in shopping_list:
                    shopping_list[ingrediente["producto"]] = {
                        ingrediente["unidad"]: ingrediente["cantidad"]
                    }

                else:
                    if ingrediente["unidad"] in shopping_list[ingrediente["producto"]]:
                        cantidad = str(ingrediente["cantidad"])
                        cantidad = float(sum(Fraction(s) for s in cantidad.split()))
                        cantidad_anterior = str(
                            shopping_list[ingrediente["producto"]][
                                ingrediente["unidad"]
                            ]
                        )
                        cantidad_anterior = float(
                            sum(Fraction(s) for s in cantidad_anterior.split())
                        )
                        cantidad += cantidad_anterior

                        shopping_list[ingrediente["producto"]][
                            ingrediente["unidad"]
                        ] = cantidad

                    else:
                        shopping_list[ingrediente["producto"]][
                            ingrediente["unidad"]
                        ] = ingrediente["cantidad"]

shopping_list_doc = Workbook()
lista_compras_sheet = shopping_list_doc.create_sheet("Lista de Compras")

lista_compras_sheet["A1"] = "Abastos"
lista_compras_sheet["B1"] = "Carniceria"
lista_compras_sheet["C1"] = "Super"
lista_compras_sheet["D1"] = "Plaza del Valle"

col = 65  # A
row_abastos = 2
row_carniceria = 2
row_super = 2
row_plaza = 2
abastos = {}
carniceria = {}
super = {}
plazaDelValle = {}
for key, value in shopping_list.items():
    quant = ""
    for key1, value1 in value.items():
        quant += f"{value1}: {key1}, "

    item = f"{key} ({quant})"
    match where_ingredients[key]:
        case "Abastos":
            lista_compras_sheet[f"{chr(col)}{row_abastos}"] = item
            abastos[key] = item
            row_abastos += 1

        case "Carniceria":
            lista_compras_sheet[f"{chr(col + 1)}{row_carniceria}"] = item
            carniceria[key] = item
            row_carniceria += 1

        case "Super":
            lista_compras_sheet[f"{chr(col + 2)}{row_super}"] = item
            super[key] = item
            row_super += 1

        case "Plaza del Valle":
            lista_compras_sheet[f"{chr(col + 3)}{row_plaza}"] = item
            plazaDelValle[key] = item
            row_plaza += 1

# Constants
lista_compras_sheet[f"{chr(col + 1)}{row_carniceria}"] = f"Patas de Pollo (1 kg)"
carniceria["Patas de Pollo"] = "Patas de Pollo (1 kg)"
row_carniceria += 1
createListsInKeep(
    {
        "Abastos": abastos,
        "Carniceria": carniceria,
        "Super": super,
        "Plaza del Valle": plazaDelValle,
    }
)

shopping_list_doc.save("/home/rammas/Downloads/shopping_list.xlsx")
